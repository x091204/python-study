import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]
## basic functions
## 1.list each company with respective product count
## 2.list product with inventory less than 10
## 3.list each company with respective total inventory value

product_per_suppliers = {}
total_value_per_supplier = {}
product_under_10_inv = {}

for product_raw in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_raw,4).value
    inventory = product_list.cell(product_raw,2).value
    price = product_list.cell(product_raw, 3).value
    product_number = product_list.cell(product_raw,1).value
    inventory_price = product_list.cell(product_raw, 5)

    ## calculation per suppliers with respective product count
    if supplier_name in product_per_suppliers:
        current_num_product = product_per_suppliers.get(supplier_name)
        product_per_suppliers[supplier_name] = current_num_product + 1
    else:
        product_per_suppliers[supplier_name] = 1

    ## calculation of total inventory value per suppliers
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory + price
    else:
        total_value_per_supplier[supplier_name] = inventory * price


    ## calculation of product that are less than 10 in inventory
    if inventory < 10:
        product_under_10_inv[int(product_number)] = int(inventory)

    ## add value for total inventory price
    inventory_price.value = inventory * price
inv_file.save("inventory.xlsx")

print(product_under_10_inv)
print(total_value_per_supplier)
print(product_per_suppliers)
